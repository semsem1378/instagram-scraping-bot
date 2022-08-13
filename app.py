from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.chrome.options import Options
import time
import re
import random
import openpyxl
from tkinter import messagebox
from progress import bar

posts = []
tags =[]
pageIds =[]
tels = []
followers =[]
websites = []

username = ''
password = ''
failed = False

# initiation :
options = Options()
options.headless = True
try:
    browser = webdriver.Firefox(options=options)
except:
    browser = webdriver.Chrome(options=options)


# login to instagram
def login(username , password):
    """
    logs into your account .
    needs username and password to be passed 
    """
    browser.get("https://www.instagram.com/")
    global failed
    try:
        exist = WebDriverWait(browser, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "input"))
            )
        # print(allCards)    
    except:
        print("login failed")
        messagebox.showwarning('vorood be account momken nist', 'username va password ra dobare check konid')
        failed = True
        return
    inputs = browser.find_elements('tag name','input')
    inputs[0].send_keys(username)
    inputs[1].send_keys(password)
    time.sleep(2)
    btn = browser.find_element('css selector', "button > div")
    print(btn)
    btn.click()
    time.sleep(4)

# logout of instagram account
def logout():
    try:
        btn = browser.find_element('css selector', f'img[alt="{username}\'s profile picture"]')
        btn.click()
        time.sleep(3)
        logout = browser.find_element( 'xpath', '//div [contains( text(), "Log Out")]')
        logout.click()
    except:
        print("could\'nt log out")

def setTag(tagname):
    """
    set a tag name to find posts with.
    """
    if tagname =="":   
        tagname = input("please enter your tag : \n\t")
    linkbyTag = f"https://www.instagram.com/explore/tags/{tagname}"
    bar(2)
    return linkbyTag


# get related posts by links
def getPosts():
    """
    finds posts with tag name
    """
    global posts
    countForThisTag =0 
    try:
        time.sleep(3)
        allLinks = browser.find_elements('css selector',"a")
        for link in allLinks:    
            temp = re.template(r"(?=\/p\/)")
            href = link.get_attribute("href")
            res = temp.findall(href)
            if res:
                posts.append(href)
                countForThisTag += 1
            
    except Exception as err:
        print(f"error in getting posts : {err}")
    time.sleep(4)
    

# write excel file 
def writer(tagname):
    try:
        wb = openpyxl.Workbook() 
        sheet = wb.active
        
        # writing ids column
        row_num = 2
        cell1 = sheet.cell(row=1, column=1)
        cell1.value = 'page id'
        for id in pageIds:
            cell = sheet.cell(row = row_num, column= 1)
            row_num += 1
            cell.value = id
        
        # writing phone numbers column 
        row_num = 2
        cell2 = sheet.cell(row=1, column=2)
        cell2.value = 'phone'
        for tel in tels:
            cell = sheet.cell(row = row_num, column= 2)
            row_num += 1
            cell.value = tel

        # writing followers column
        row_num = 2
        cell3 = sheet.cell(row=1, column=3)
        cell3.value = 'followers'
        for follow in followers:
            cell = sheet.cell(row = row_num, column= 3)
            row_num += 1
            cell.value = follow
        
        # writing website column
        row_num = 2
        cell4 = sheet.cell(row=1, column=4)
        cell4.value = 'website'
        for web in websites:
            cell = sheet.cell(row = row_num, column= 4)
            row_num += 1
            cell.value = web
            
        wb.save(filename=f'{tagname}{random.randint(0,358)}.xlsx')
    except :
        print('failed to save proper result')
    bar(2)
            

# scroll to find more posts
def scroller(max_amount):
    last_height = browser.execute_script("return document.body.scrollHeight")
    eachPers = (1/max_amount) * 26
    while True: 
        getPosts()
        bar(eachPers)
        print(f' posts: {len(posts)}')
        # Scroll down to bottom
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(8)
        
        #Get new height of screen
        new_height = browser.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            print(len(posts))
            break
        last_height = new_height
        if(len(posts) > max_amount):
            print(f'reached {max_amount} !')
            break
   

# processing data 
def getData():
    global tels, pageIds, posts
    
    # setting percentage of progress bar by each data row recieved
    eachPers = (1/len(posts)) * 70
    for link in posts:
        browser.get(link)
        bar(eachPers)
        try:
            header = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "header._aaqw")))
        except:
            print('could not find header', link + '\n****\n')
            continue
        try: 
            div = header.find_element('css selector', 'div')
            id = div.find_element('css selector', 'a').get_attribute("href")
        except Exception as err:
            print('div or a not found', f'{err} \n****\n')
            continue
        id = id[:len(id)-1].replace('https://www.instagram.com/', '')
        if id in pageIds:
            continue
        else:
            pageIds.append(id)
        browser.get(f'https://www.instagram.com/{id}/')
        try:
            bio = WebDriverWait(browser, 10).until( EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div._aacl._aacp._aacu._aacx._aad6._aade')))
            txt = bio[3].get_attribute("innerHTML")
            x = re.search(r"\d{3}\d{3}\d{5}", txt)
            
        except Exception as err:
            print('no bio available', f'{err} \n****\n')
            tels.append('not found')
            followers.append('')
            websites.append('')
            continue
        new_string = None
        if x:
            print(x.group())
            tels.append(x.group())
            print('added')
        elif re.search(r'تماس',txt) or re.search(r'تلفن',txt):
            new_string = ''.join(re.findall(r'\d+', txt))
        else:
            tels.append('not found')
        if new_string:
            print(new_string)
            if new_string[0] == '0' or new_string[0] == '۰':
                tels.append(new_string[:12])
                print('added')
            elif len(new_string) == 10:
                tels.append(f'0{new_string}')
                print('added')
            else:
                tels.append('not accessable')       
        follow = bio[1].find_element('css selector', 'span').get_attribute('innerHTML')
        followers.append(follow)
        try:
            a = header.find_element('css selector', 'a')
            website = a.find_element('css selector', 'div').get_attribute('innerHTML')
            websites.append(website)
        except:
            websites.append('none')
        time.sleep(3)       
        
  
# main app      
def main(user_name, pass_word, posts_number, hash_tag):
    
    #setting variables
    global username
    username = user_name
    hash_tag = hash_tag.decode('utf-8')
    print('here', f'{user_name}, {pass_word},{hash_tag},{posts_number}')
    
    login(username=user_name, password=pass_word)

    # checking login attemp
    if failed:
        # endApp()
        return
    
    link = setTag(hash_tag)
    browser.get(link)
    
    # scrolling page to get all posts needed
    scroller(max_amount=posts_number)
    print('all post are : \n',len(posts))
    
    # checking whether any posts found or not
    if len(posts) == 0:
        messagebox.showerror('failed', 'posti yaft nashod ! lotfan dobare talash konid')
        # endApp()
        return
    
    # getting account info from selected posts
    getData()
    print('all accounts are : \n',len(pageIds))
    
    # writing all info to a .xlsx file
    writer(str(hash_tag))
    
    # logging out from account
    logout()
    messagebox.showinfo('Done', 'All Done !')
    




